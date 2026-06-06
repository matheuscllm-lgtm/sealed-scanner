#!/usr/bin/env python3
"""run_all_sources.py — orquestrador multi-fonte do TCG Sealed Arbitrage Scanner.

Roda as 3 fontes BR (Amazon, Liga, OLX) numa execução e consolida TUDO numa
tabela única interpretável (`unified_deals.csv` + `unified_deals.xlsx`), com
coluna Fonte e ordenada por bucket (GREEN → YELLOW → RED) e margem desc.

Comportamento robusto (autonomia / keep-alive):
- OLX bloqueado por Cloudflare WAF (SourceBlockedError) = condição EXTERNA
  não-fatal: registra "bloqueada", segue com as outras fontes.
- Falha transitória de uma fonte = 1 retry; se persistir, registra "falhou"
  e segue (não derruba o run inteiro).
- Só falha (exit 1) se NENHUMA fonte entregou anúncios — aí algo real quebrou.
- Cada fonte logada com status, contagem e tempo. Tudo auditável.

Uso:
    python run_all_sources.py                 # amazon + liga + olx
    python run_all_sources.py --sources amazon,liga
"""
from __future__ import annotations

import argparse
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import sealed_arbitrage_scanner as s
from lib.errors import SourceBlockedError

DEFAULT_SOURCES = ["amazon", "liga", "olx", "mercadolivre"]  # liga é o longo; olx/ml via firecrawl
BUCKET_ORDER = {"real_opportunities": 0, "review_required": 1, "rejected": 2}


def _build_rows(listings: list[dict], source: str, registry, us_reference, config) -> list:
    """Replica o loop de classificação do scanner.run() para uma fonte."""
    rows = []
    for item in listings:
        qty_raw = item.get("qty_avail")
        try:
            qty_avail = int(qty_raw) if qty_raw is not None else None
        except (TypeError, ValueError):
            qty_avail = None
        row = s.ScanRow(
            listing_id=str(item.get("id", "")),
            title_br=item.get("title", ""),
            source=item.get("source", source),
            seller=item.get("seller", ""),
            url=item.get("url", ""),
            price_brl=float(item.get("price_brl", 0.0)),
            qty_avail=qty_avail,
        )
        rows.append(s.classify(row, registry, us_reference, config))
    return rows


def _scan_one(source: str, config, registry, registry_raw, us_reference,
              mock_path: Path, retries: int = 1) -> dict:
    """Roda uma fonte. Retorna dict com status/rows/erro. Nunca levanta."""
    attempt = 0
    while True:
        attempt += 1
        t0 = time.time()
        try:
            listings, desc = s.load_listings(source, mock_path, config, registry_raw=registry_raw)
            rows = _build_rows(listings, source, registry, us_reference, config)
            return {
                "source": source, "status": "ok", "rows": rows,
                "n_listings": len(listings), "desc": desc,
                "elapsed_s": round(time.time() - t0, 1), "error": "",
            }
        except SourceBlockedError as exc:
            return {
                "source": source, "status": "blocked", "rows": [],
                "n_listings": 0, "desc": str(exc),
                "elapsed_s": round(time.time() - t0, 1), "error": str(exc),
            }
        except Exception as exc:  # transitório — retry
            err = f"{type(exc).__name__}: {exc}"
            print(f"  [orq] fonte {source} falhou (tentativa {attempt}): {err}")
            traceback.print_exc()
            if attempt > retries:
                return {
                    "source": source, "status": "failed", "rows": [],
                    "n_listings": 0, "desc": "",
                    "elapsed_s": round(time.time() - t0, 1), "error": err,
                }
            time.sleep(3)


def _write_unified(all_rows: list, summaries: list[dict], out_dir: Path, stamp: str) -> Path | None:
    """Escreve unified_deals.csv + .xlsx com todas as fontes numa tabela só."""
    # Ordena: bucket (GREEN→YELLOW→RED), depois margem total desc.
    def sort_key(r):
        b = BUCKET_ORDER.get(r.bucket, 3)
        m = r.total_margin_pct if r.total_margin_pct is not None else -999
        return (b, -m)

    all_rows = sorted(all_rows, key=sort_key)

    csv_path = out_dir / "unified_deals.csv"
    s.write_csv(all_rows, csv_path)

    xlsx_path = out_dir / f"unified_sealed_{stamp}.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Todos os Deals"
        ws.append([label for _, label in s.CSV_COLUMNS])
        for c in ws[1]:
            c.font = Font(bold=True)
        fills = {
            "real_opportunities": PatternFill("solid", fgColor="C6EFCE"),
            "review_required": PatternFill("solid", fgColor="FFEB9C"),
            "rejected": PatternFill("solid", fgColor="F2F2F2"),
        }
        for row in all_rows:
            ws.append([s.cell_value(row, key) for key, _ in s.CSV_COLUMNS])
            fill = fills.get(row.bucket)
            if fill:
                for c in ws[ws.max_row]:
                    c.fill = fill
        # larguras
        for idx, (key, label) in enumerate(s.CSV_COLUMNS, start=1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(46, len(label) + 4))
        ws.freeze_panes = "A2"

        # aba Resumo por fonte
        ws2 = wb.create_sheet("Resumo")
        ws2.append(["Fonte", "Status", "Anúncios", "GREEN", "YELLOW", "RED", "Tempo (s)", "Detalhe"])
        for c in ws2[1]:
            c.font = Font(bold=True)
        for info in summaries:
            ws2.append([
                info["source"], info["status"], info["n_listings"],
                info["green"], info["yellow"], info["red"],
                info["elapsed_s"], (info["desc"] or info["error"])[:80],
            ])
        wb.save(xlsx_path)
        return xlsx_path
    except Exception as exc:
        print(f"  [orq] aviso: XLSX falhou ({exc}); CSV gerado em {csv_path}")
        return None


def run(args: argparse.Namespace) -> int:
    config = s.load_yaml(Path(args.config), "config.yaml")
    fx_source = s.resolve_fx_rate(config)
    config["currency"]["_source"] = fx_source
    print(f"  [fx] cambio USD/BRL: {fx_source}")
    registry_data = s.load_yaml(Path(args.registry), "sku_registry.yaml")
    registry = s.build_registry(registry_data)
    registry_raw = registry_data.get("skus", [])
    ref_data = s.load_json(SCRIPT_DIR / "data" / "us_reference.json", "data/us_reference.json")
    us_reference = ref_data.get("prices", {})
    mock_path = Path(args.mock)

    sources = [x.strip() for x in args.sources.split(",") if x.strip()]
    print(f"  [orq] fontes: {sources}")

    all_rows: list = []
    summaries: list[dict] = []
    for source in sources:
        print(f"\n  [orq] >>> rodando fonte: {source}")
        info = _scan_one(source, config, registry, registry_raw, us_reference, mock_path)
        rows = info.pop("rows")
        info["green"] = sum(1 for r in rows if r.bucket == "real_opportunities")
        info["yellow"] = sum(1 for r in rows if r.bucket == "review_required")
        info["red"] = sum(1 for r in rows if r.bucket == "rejected")
        summaries.append(info)
        all_rows.extend(rows)
        print(f"  [orq] <<< {source}: status={info['status']} "
              f"anúncios={info['n_listings']} "
              f"(GREEN={info['green']} YELLOW={info['yellow']} RED={info['red']}) "
              f"{info['elapsed_s']}s")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = SCRIPT_DIR / "results" / f"unified_{stamp}"
    out_dir.mkdir(parents=True, exist_ok=True)
    # Marcador parseável pelo watchdog (grava run.log no dir CERTO, sem
    # heurística de "dir mais novo" que erra sob escrita concorrente).
    print(f"UNIFIED_OUT_DIR={out_dir}")

    xlsx_path = _write_unified(all_rows, summaries, out_dir, stamp)

    # Banner final
    green = sum(s_["green"] for s_ in summaries)
    yellow = sum(s_["yellow"] for s_ in summaries)
    red = sum(s_["red"] for s_ in summaries)
    print("\n" + "=" * 64)
    print("  TCG SEALED — SCAN UNIFICADO (Amazon + Liga + OLX)")
    print("=" * 64)
    print(f"  Câmbio USD/BRL : {config['currency']['usd_brl']:.4f}  [{fx_source}]")
    for info in summaries:
        tag = {"ok": "OK", "blocked": "BLOQUEADA", "failed": "FALHOU"}.get(info["status"], info["status"])
        print(f"  {info['source']:<8} {tag:<10} anúncios={info['n_listings']:<5} "
              f"GREEN={info['green']} YELLOW={info['yellow']} RED={info['red']}  ({info['elapsed_s']}s)")
    print("-" * 64)
    print(f"  TOTAL  GREEN={green}  YELLOW={yellow}  RED={red}  (linhas={len(all_rows)})")
    print(f"  CSV   : {out_dir / 'unified_deals.csv'}")
    if xlsx_path:
        print(f"  XLSX  : {xlsx_path}")
    print("=" * 64)

    # exit 1 só se NENHUMA fonte entregou nada (quebra real, não block)
    delivered = any(info["status"] == "ok" and info["n_listings"] > 0 for info in summaries)
    any_real_failure = any(info["status"] == "failed" for info in summaries)
    if not delivered:
        print("  ERRO: nenhuma fonte entregou anúncios.")
        return 1
    if any_real_failure:
        print("  AVISO: ao menos uma fonte falhou (transitório); tabela parcial entregue.")
        return 0
    return 0


def main() -> None:
    from lib.console import harden_stdout
    harden_stdout()  # console Windows cp1252 quebra em títulos Liga/PT-BR
    p = argparse.ArgumentParser(description="Scan unificado multi-fonte (sealed BR -> US)")
    p.add_argument("--sources", default=",".join(DEFAULT_SOURCES),
                   help="fontes separadas por vírgula (default: amazon,liga,olx,mercadolivre)")
    p.add_argument("--config", default=str(SCRIPT_DIR / "config.yaml"))
    p.add_argument("--registry", default=str(SCRIPT_DIR / "sku_registry.yaml"))
    p.add_argument("--mock", default=str(SCRIPT_DIR / "mock_data" / "liga_listings.json"))
    args = p.parse_args()
    sys.exit(run(args))


if __name__ == "__main__":
    main()
