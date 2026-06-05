#!/usr/bin/env python3
"""build_delivery_xlsx.py — XLSX condensado p/ entrega (Drive via MCP).

Lê o unified_deals.csv de um scan e gera um XLSX pequeno só com os deals
ACIONÁVEIS (GREEN + YELLOW) + aba Resumo. O arquivo cheio (~800 linhas, a
maioria RED) é grande demais pra upload inline do MCP; este condensado cabe.

Uso: python scripts/build_delivery_xlsx.py [<dir do scan>] [<saida.xlsx>]
     (sem args: usa o results/unified_* mais recente; saida em TEMP)
"""
from __future__ import annotations

import csv
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent

COLS = [
    ("Confiança do deal", "Tier"),
    ("Fonte", "Fonte"),
    ("Produto (canônico)", "Produto"),
    ("Tipo", "Tipo"),
    ("Preço BR (R$)", "BR R$"),
    ("Qtd disponível", "Qtd"),     # estoque do vendedor — importamos em LOTE, nunca 1 unid
    ("Preço US (R$)", "US R$"),
    ("Margem total %", "Margem %"),
    # Margem líquida REMOVIDA (operador 2026-06-02): sem saber frete real +
    # tamanho do lote por remessa, o líquido é um número fabricado. Não exibir.
    ("Confiança do match", "Match"),
    ("URL", "URL"),
]
FILLS = {
    "GREEN": PatternFill("solid", fgColor="C6EFCE"),
    "YELLOW": PatternFill("solid", fgColor="FFEB9C"),
}
ORDER = {"GREEN": 0, "YELLOW": 1}

sys.path.insert(0, str(ROOT))
try:
    from pool_fill import avg_price_for_sku
except Exception:
    avg_price_for_sku = None


def latest_scan_dir() -> Path:
    dirs = sorted((ROOT / "results").glob("unified_*"), key=lambda d: d.stat().st_mtime, reverse=True)
    if not dirs:
        sys.exit("ERRO: nenhum results/unified_* encontrado.")
    return dirs[0]


def _margin(row: dict) -> float:
    try:
        return float(row.get("Margem total %") or -999)
    except ValueError:
        return -999.0


def _int_or_none(v: str | None) -> int | None:
    try:
        return int(float(v)) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def main() -> None:
    scan_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else latest_scan_dir()
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(tempfile.gettempdir()) / "sealed_delivery.xlsx"
    src = scan_dir / "unified_deals.csv"
    rows = list(csv.DictReader(src.open(encoding="utf-8")))
    sel = [r for r in rows if r["Confiança do deal"] in ("GREEN", "YELLOW")]
    sel.sort(key=lambda r: (ORDER.get(r["Confiança do deal"], 9), -_margin(r)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Deals (GREEN+YELLOW)"
    ws.append([lbl for _, lbl in COLS])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in sel:
        ws.append([r.get(k, "") for k, _ in COLS])
        fill = FILLS.get(r["Confiança do deal"])
        if fill:
            for c in ws[ws.max_row]:
                c.fill = fill
    for i, (_, lbl) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(44, len(lbl) + 4))
    ws.freeze_panes = "A2"

    green = sum(1 for r in sel if r["Confiança do deal"] == "GREEN")
    yellow = len(sel) - green
    ws2 = wb.create_sheet("Resumo")
    summary = [
        ("Scan", scan_dir.name),
        ("GREEN", green),
        ("YELLOW", yellow),
        ("RED (não incluídos)", len(rows) - len(sel)),
        ("Total linhas scan", len(rows)),
    ]
    for k, v in summary:
        ws2.append([k, v])
    for i in range(1, len(summary) + 1):
        ws2[f"A{i}"].font = Font(bold=True)

    # aba Preço médio por SKU — custo real ao varrer vários logistas (sem frete)
    if avg_price_for_sku is not None:
        by_sku: dict[str, list[dict]] = {}
        for r in sel:
            sku = (r.get("SKU") or "").strip()
            if not sku or r.get("Confiança do match") != "HIGH":
                continue
            by_sku.setdefault(sku, []).append(r)
        avg_rows = []
        for sku, rs in by_sku.items():
            listings = []
            for r in rs:
                try:
                    price = float(r.get("Preço BR (R$)") or "")
                except ValueError:
                    continue
                listings.append({"seller": r.get("Vendedor", ""), "price_brl": price,
                                 "qty_avail": _int_or_none(r.get("Qtd disponível"))})
            if not listings:
                continue
            try:
                us_brl = float(rs[0].get("Preço US (R$)") or 0)
            except ValueError:
                us_brl = 0.0
            avg_rows.append((rs[0], avg_price_for_sku(listings, us_price_brl=us_brl, sku=sku)))
        if avg_rows:
            avg_rows.sort(key=lambda t: -t[1].avg_margin_pct)
            ws3 = wb.create_sheet("Preço médio")
            ws3.append(["Produto", "Tipo", "Tier", "# Vend.", "Qtd total",
                        "Melhor R$", "Médio R$", "US R$", "Margem média %"])
            for c in ws3[1]:
                c.font = Font(bold=True)
            for first, a in avg_rows:
                tier = first.get("Confiança do deal", "")
                ws3.append([first.get("Produto (canônico)", ""), first.get("Tipo", ""), tier,
                            a.n_sellers, a.total_qty, round(a.best_price, 2),
                            round(a.weighted_avg_price, 2), round(a.us_price_brl, 2),
                            round(a.avg_margin_pct, 1)])
                fill = FILLS.get(tier)
                if fill:
                    for c in ws3[ws3.max_row]:
                        c.fill = fill
            for i, w in enumerate([40, 16, 8, 8, 10, 12, 12, 12, 14], 1):
                ws3.column_dimensions[get_column_letter(i)].width = w
            ws3.freeze_panes = "A2"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"OK | GREEN={green} YELLOW={yellow} | {out} | {out.stat().st_size} bytes")
    print(f"PATH={out}")


if __name__ == "__main__":
    main()
