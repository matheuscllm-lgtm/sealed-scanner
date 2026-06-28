#!/usr/bin/env python3
"""
TCG Sealed Arbitrage Scanner

Brasil (Liga Pokémon) -> EUA (TCGPlayer). Produtos selados Pokémon TCG.

Pergunta que o scanner responde:
  "Esse produto selado comprado no Brasil, por esse preço, ainda dá lucro
   se revendido nos EUA pelo preço do TCGPlayer, depois de taxas conservadoras?"

PRIMEIRA VERSÃO — mock-first
  O pipeline completo (carregar anúncios -> casar com SKU -> calcular margem
  -> classificar -> relatórios) roda sobre dados MOCKADOS. O adapter da Liga
  Pokémon (`--source liga`, em liga_adapter.py) é um esboço: o fetch via
  patchright já está pronto, mas o parsing das páginas de selado aguarda a
  investigação com probe_liga_sealed.py (a Liga responde 403 a IP de
  datacenter; o mapeamento roda local, como os probe_liga_*.py da raiz).

Uso:
  python sealed_arbitrage_scanner.py                 # roda sobre mock_data/
  python sealed_arbitrage_scanner.py --source mock
  python sealed_arbitrage_scanner.py --source liga   # NotImplementedError

Saída:
  sealed/results/<timestamp>/
    real_opportunities.csv   oportunidades GREEN
    review_required.csv      YELLOW + matches ambíguos
    rejected.csv             sem match / sem referência / margem ruim
    sealed_scan_<timestamp>.xlsx   tudo acima + premissas + summary

Princípios:
  - Nunca inventar preço. Sem referência US -> rejeitado, não estimado.
  - Cada scan é fresco: diretório novo por execução, nada de misturar runs.
  - Premissas (câmbio, taxas) sempre impressas no relatório.

Exit codes:
  0  run saudável (com ou sem oportunidades)
  1  nenhum anúncio carregado (fonte provavelmente quebrada)
  2  erro de configuração (arquivo faltando / inválido)
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

try:
    import yaml
except ImportError:
    print("ERRO: PyYAML não instalado. Rode: pip install -r requirements.txt")
    sys.exit(2)

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from lib.errors import SourceBlockedError

# Idiomas não-ingleses que invalidam o match (vendemos EN no TCGPlayer).
NON_EN_LANGUAGE_TOKENS = ("japones", "japonesa", "coreano", "coreana", "chines", "chinesa")
# Padrão de numeração de carta avulsa, ex.: "238/191".
CARD_NUMBER_RE = re.compile(r"\b\d{1,3}\s*/\s*\d{1,3}\b")


# --------------------------------------------------------------------------
# Normalização
# --------------------------------------------------------------------------
def normalize(text: str) -> str:
    """minúsculas, acentos removidos, pontuação -> espaço, espaços colapsados.

    "Booster Box Surging Sparks (Inglês)" -> "booster box surging sparks ingles"
    """
    if not text:
        return ""
    decomposed = unicodedata.normalize("NFKD", str(text))
    stripped = "".join(c for c in decomposed if not unicodedata.combining(c))
    cleaned = "".join(c if c.isalnum() else " " for c in stripped.lower())
    return " ".join(cleaned.split())


def contains_term(haystack_norm: str, term: str) -> bool:
    """Match de termo por palavra inteira (evita 'tin' casar dentro de 'trainer')."""
    return f" {normalize(term)} " in f" {haystack_norm} "


# Sinais FORTES de carta avulsa (single) — fora do escopo de selado. Um single
# cujo título cita o set + o tipo do box de origem (ex.: "Eevee SVP 173 –
# Prismatic Evolutions – Elite Trainer Box Promo - Single Card") casava com o
# SKU do box e gerava MARGEM FANTASMA (preço da carta vs preço do box → +272%).
# Estes sinais são impossíveis em produto selado legítimo, então um título que
# os exibe NÃO deve casar com nenhum SKU (rejeitado antes do match).
# NB: "single" sozinho é proibido aqui — existe "Sealed Single Booster Pack".
_SINGLE_CARD_TOKENS = (
    "single card",          # EN explícito (caso Eevee)
    "carta avulsa", "cartas avulsas", "carta unica", "cartas unicas",  # PT
)
# "SVP <n>" = Scarlet & Violet Promo: numeração de carta promo avulsa, nunca selado.
_PROMO_SINGLE_RE = re.compile(r"\bsvp\s*\d+\b")


def looks_like_single_card(title: str) -> bool:
    """True se o título tem sinal forte de carta avulsa (single), fora do escopo.

    Cobre: 'single card' (EN), 'carta avulsa/única' (PT), código de promo
    'SVP 173', e numeração de carta 'NNN/NNN' (CARD_NUMBER_RE). Viés conservador
    de scanner de COMPRA: na dúvida, rejeitar (falso-negativo = perde 1 deal;
    falso-positivo = deal fantasma que custa tempo/risco ao operador)."""
    norm = normalize(title)
    if any(tok in norm for tok in _SINGLE_CARD_TOKENS):
        return True
    if _PROMO_SINGLE_RE.search(norm):
        return True
    if CARD_NUMBER_RE.search(title):  # 'NNN/NNN' — a barra some na normalização
        return True
    return False


# Sinais de ACESSÓRIO puro (porta-cartas, playmat, toploader...) — fora do escopo
# de selado. Caso real (scan 2026-06-06, MercadoLivre): "Elite Trainer Box -
# Ascended Heroes - Acessórios" (R$95) casava ah-etb-en → +885% fantasma.
#
# ⚠️ TOKENS PROIBIDOS aqui (colidem com SKU selado curado — checado nos
# match-terms do registry 2026-06-06, NÃO só nos nomes):
#   - 'binder'/'fichario'/'album' → os 4 "Collection Box" (blk/wht/pre/mew) têm
#     type_term 'binder collection'; suas listagens BR usam "fichário"/"álbum".
#     Rejeitá-los barraria produto selado real.
#   - 'sleeve' → 'Sleeved Booster' = 20 SKUs selados.
#   - 'collection'/'box' → caixas seladas reais.
# (O "Fichário Binder" +440% do scan NÃO é acessório nem bug de dado: auditoria
#  2026-06-09 via tcgcsv confirmou que 502004 = "151 Binder Collection" e o
#  market US$241 é real (low US$236; produto out-of-print valorizado). Margens
#  >200% caem no guard `margem_anomala` p/ verificação manual do anúncio.)
_ACCESSORY_TOKENS = (
    "acessorio", "acessorios",   # acessório vendido à parte (viés conservador)
    "porta carta", "porta cartas", "porta card", "porta deck",
    "toploader", "top loader", "playmat", "tapete", "deck shield",
    "deck protector", "deck protetor",   # protetor de cartas avulso
)
# Pacote de SLEEVES (protetores de carta) vendido à parte — acessório, não selado.
# Caso real (scan 2026-06-26, MercadoLivre): "Sleeve Dragonite Etb Ascended Heroes
# (65 Sleeves)" R$64,50 casava ah-etb-en e dava +1600% (o guard margem_anomala
# pegava como RED, mas a oferta envenenava a entrega do grupo — virava a oferta de
# referência e escondia o ETB GREEN real). 'sleeve' nu é PROIBIDO aqui (colide com
# 'Sleeved Booster', 20 SKUs selados); um CONTADOR de sleeves ("65 Sleeves") é o
# sinal seguro: nenhum selado real diz "N sleeves" (o Sleeved Booster é 1 pacote, e
# normaliza pra "sleeved", nunca "N sleeves").
_SLEEVE_PACK_RE = re.compile(r"\b\d+\s*sleeves?\b")


def looks_like_accessory(title: str) -> bool:
    """True se o título tem sinal forte de ACESSÓRIO puro (não-selado), fora do escopo.

    Mesmo viés conservador de scanner de COMPRA do `looks_like_single_card`.
    NÃO usa 'sleeve'/'binder'/'fichario'/'album'/'collection'/'box' (singular) —
    todos são (parte de) produtos selados reais neste registry. MAS um contador de
    sleeves ("65 sleeves") é pacote de protetores avulso = acessório."""
    norm = normalize(title)
    if _SLEEVE_PACK_RE.search(norm):
        return True
    return any(tok in norm for tok in _ACCESSORY_TOKENS)


# Sinais EXPLÍCITOS de produto ABERTO/USADO/INCOMPLETO — nunca é uma compra de
# selado. O scanner é SELADO-only (`scope.exclude`: Opened/Damaged), mas até
# 2026-06-21 NADA no código barrava isso — só funcionava porque Liga/OLX/ML são
# "new-first". Um box aberto/sem cartas casado a um SKU selado = margem fantasma.
# Rejeição GLOBAL (todas as fontes): validado zero-regressão (0 de 818 matches
# reais têm qualquer destes tokens). Viés conservador de COMPRA: na dúvida, barrar.
_USED_TOKENS = (
    "usado", "usada", "usados", "usadas", "aberto", "aberta", "abertos", "abertas",
    "sem cartas", "sem lacre", "sem o lacre", "sem plastico", "sem o plastico",
    "sem booster", "sem boosters", "sem os boosters", "so a caixa", "somente a caixa",
    "apenas a caixa", "vazio", "vazia", "incompleto", "incompleta",
    "destrocado", "destrocada", "danificado", "danificada",
)
# Sinais EXPLÍCITOS de LACRE — prova positiva de selado. Exigidos só para fontes
# `sealed_only` (marketplaces secondhand-first, ex.: Enjoei), onde o default tem
# de ser "usado até provar lacre". Fontes new-first NÃO exigem (Liga/OLX/ML novos
# raramente escrevem "lacrado") — manter o default inalterado = zero regressão.
_SEALED_TOKENS = (
    "lacrado", "lacrada", "lacrados", "lacradas", "selado", "selada", "selados",
    "seladas", "sealed", "factory sealed", "novo lacrado", "nova lacrada",
    "lacre", "lacre de fabrica", "com lacre",   # substantivo (NB: "sem lacre" cai antes no gate de usado)
)


def looks_used(title: str) -> bool:
    """True se o título sinaliza produto ABERTO/USADO/INCOMPLETO (fora do escopo)."""
    norm = normalize(title)
    return any(contains_term(norm, tok) for tok in _USED_TOKENS)


def looks_sealed(title: str) -> bool:
    """True se o título tem prova POSITIVA de lacre (lacrado/selado/sealed)."""
    norm = normalize(title)
    return any(contains_term(norm, tok) for tok in _SEALED_TOKENS)


# --------------------------------------------------------------------------
# Config e registry
# --------------------------------------------------------------------------
def load_yaml(path: Path, label: str) -> dict:
    if not path.exists():
        print(f"ERRO: {label} não encontrado em {path}")
        sys.exit(2)
    try:
        with path.open(encoding="utf-8") as fh:
            data = yaml.safe_load(fh)
    except yaml.YAMLError as exc:
        print(f"ERRO: {label} inválido ({path}): {exc}")
        sys.exit(2)
    if not isinstance(data, dict):
        print(f"ERRO: {label} ({path}) não é um mapeamento YAML.")
        sys.exit(2)
    return data


# --------------------------------------------------------------------------
# Câmbio USD->BRL — fetch automático com fallback
# --------------------------------------------------------------------------
def fetch_usd_brl() -> tuple[float | None, str]:
    """Busca cotação USD->BRL em fontes públicas (sem auth, free).

    Retorna (rate, source). rate=None se todas as fontes falharem.
    Tenta AwesomeAPI (BR, bid bancário) e depois open.er-api.com (backup).
    """
    import urllib.request
    import urllib.error

    sources = [
        (
            "AwesomeAPI",
            "https://economia.awesomeapi.com.br/last/USD-BRL",
            lambda d: float(d["USDBRL"]["bid"]),
        ),
        (
            "open.er-api.com",
            "https://open.er-api.com/v6/latest/USD",
            lambda d: float(d["rates"]["BRL"]),
        ),
    ]
    for name, url, extract in sources:
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "sealed-scanner/1.0"})
            with urllib.request.urlopen(req, timeout=5) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
            rate = extract(payload)
            if rate and rate > 0:
                return rate, name
        except (urllib.error.URLError, OSError, ValueError, KeyError, TypeError):
            continue
    return None, "todas as fontes falharam"


def resolve_fx_rate(config: dict) -> str:
    """Resolve cambio USD->BRL conforme config.currency.mode.

    mode=manual: usa config.currency.usd_brl direto.
    mode=fetch : tenta APIs publicas; se falhar, cai pra usd_brl manual.

    Mutates config["currency"]["usd_brl"] in-place quando fetch sucede.
    Retorna string descritiva pra log/relatorio.
    """
    currency = config.setdefault("currency", {})
    mode = (currency.get("mode") or "manual").lower()
    manual_value = currency.get("usd_brl")

    if mode != "fetch":
        return f"manual ({manual_value})"

    rate, source = fetch_usd_brl()
    if rate is None:
        print(f"  [fx] AVISO: fetch automatico falhou ({source}); usando fallback manual {manual_value}")
        return f"manual fallback ({manual_value}); fetch falhou"

    currency["usd_brl"] = rate
    return f"fetch via {source} ({rate:.4f})"


def load_json(path: Path, label: str) -> dict:
    if not path.exists():
        print(f"ERRO: {label} não encontrado em {path}")
        sys.exit(2)
    try:
        with path.open(encoding="utf-8") as fh:
            return json.load(fh)
    except json.JSONDecodeError as exc:
        print(f"ERRO: {label} inválido ({path}): {exc}")
        sys.exit(2)


# --------------------------------------------------------------------------
# Modelos
# --------------------------------------------------------------------------
@dataclass
class Sku:
    id: str
    name: str
    product_type: str
    set_name: str
    language: str
    set_terms: list[str]
    type_terms: list[str]
    exclude_terms: list[str]
    # Termos OBRIGATÓRIOS (todos precisam estar no título). Útil pra variantes
    # tipo Pokemon Center ETB — sem isso, ela colide com a ETB regular.
    requires_terms: list[str] = field(default_factory=list)
    # Quantidade típica de compra em lote (packs ≥24, boxes/ETBs = 1). Metadado do
    # registry; NÃO entra na classificação (custos operacionais ficam fora do scanner).
    bulk_qty: int = 1
    # SKU "guarda-chuva de era": o set é o nome de uma ERA que prefixa títulos
    # dos sub-sets (ex.: "Mega Evolution" prefixa Perfect Order, Chaos Rising,
    # Ascended Heroes, Phantasmal Flames...). Quando um anúncio casa este SKU E
    # um SKU de sub-set (não-umbrella), o umbrella perde — o sub-set é mais
    # específico. Resolve o over-match da era genericamente, sem enumerar
    # sub-sets em exclude_terms (que vira whack-a-mole a cada set novo).
    era_umbrella: bool = False


@dataclass
class ScanRow:
    listing_id: str
    title_br: str
    source: str
    seller: str
    url: str
    price_brl: float
    qty_avail: int | None = None      # F1.5+: estoque disponível por vendedor (None se adapter não parseou)
    sku_id: str = ""
    sku_name: str = ""
    product_type: str = ""
    set_name: str = ""
    match_confidence: str = ""        # HIGH / REVIEW / NONE
    match_candidates: list[str] = field(default_factory=list)
    us_price_usd: float | None = None
    usd_brl: float | None = None
    us_price_brl: float | None = None
    gross_profit_brl: float | None = None
    total_margin_pct: float | None = None      # (US - BR) / BR  -> filtro principal
    us_discount_pct: float | None = None        # (US - BR) / US  -> "mais barato que US"
    deal_confidence: str = ""         # GREEN / YELLOW / RED
    bucket: str = ""                  # real_opportunities / review_required / rejected
    main_risk: str = ""
    recommended_action: str = ""
    reject_reason: str = ""


# --------------------------------------------------------------------------
# Carregamento de anúncios (adapters)
# --------------------------------------------------------------------------
def load_listings(source: str, mock_path: Path, config: dict,
                  registry_raw: list[dict] | None = None) -> tuple[list[dict], str]:
    """Retorna (listings, descrição_da_fonte)."""
    if source == "mock":
        data = load_json(mock_path, "mock liga_listings.json")
        listings = data.get("listings", [])
        return listings, f"mock ({mock_path.name}, scanned_at={data.get('scanned_at', '?')})"

    if source == "liga":
        # Liga Pokémon via ScraperAPI (env SCRAPERAPI_KEY). Itera as
        # categorias em config.liga.categorias, baixa cada produto e
        # decodifica os preços anti-scraping via template matching.
        import liga_adapter
        listings = liga_adapter.fetch_listings(config)
        return listings, f"liga (ligapokemon.com.br via ScraperAPI, {len(listings)} listagens)"

    if source == "amazon":
        # Amazon BR: API pública acessível, busca por SKU. Sem CF, sem auth.
        import amazon_adapter
        listings = amazon_adapter.fetch_listings(config, registry_raw or [])
        return listings, f"amazon (amazon.com.br — busca ao vivo, {len(listings)} listagens)"

    if source == "olx":
        # OLX BR: __NEXT_DATA__ JSON embutido na busca, sem CF, sem auth.
        # Predomina PT-BR; o matcher filtra via exclude_terms.
        import olx_adapter
        listings = olx_adapter.fetch_listings(config, registry_raw or [])
        return listings, f"olx (olx.com.br/brasil — busca ao vivo, {len(listings)} listagens)"

    if source == "mercadolivre":
        # Mercado Livre BR: parser DOM/CSS (sem __NEXT_DATA__). Anti-bot próprio
        # do ML → firecrawl-first (proxy stealth + waitFor ~14s). Predomina
        # PT-BR/COPAG; o matcher filtra via exclude_terms.
        import mercadolivre_adapter
        listings = mercadolivre_adapter.fetch_listings(config, registry_raw or [])
        return listings, f"mercadolivre (mercadolivre.com.br — busca ao vivo, {len(listings)} listagens)"

    print(f"ERRO: fonte desconhecida '{source}'. Use 'mock', 'amazon', 'olx', 'mercadolivre' ou 'liga'.")
    sys.exit(2)


# --------------------------------------------------------------------------
# Matcher — registry curado
# --------------------------------------------------------------------------
def build_registry(registry_data: dict) -> list[Sku]:
    skus: list[Sku] = []
    for entry in registry_data.get("skus", []):
        match = entry.get("match", {})
        skus.append(Sku(
            id=entry["id"],
            name=entry.get("name", entry["id"]),
            product_type=entry.get("product_type", ""),
            set_name=entry.get("set", ""),
            language=entry.get("language", ""),
            set_terms=match.get("set_terms", []),
            type_terms=match.get("type_terms", []),
            exclude_terms=match.get("exclude_terms", []),
            requires_terms=match.get("requires_terms", []),
            bulk_qty=int(entry.get("bulk_qty", 1)),
            era_umbrella=bool(entry.get("era_umbrella", False)),
        ))
    if not skus:
        print("ERRO: sku_registry.yaml não tem nenhum SKU.")
        sys.exit(2)
    return skus


def match_listing(title: str, registry: list[Sku], sealed_only: bool = False) -> list[Sku]:
    """SKUs candidatos: set_term casa E type_term casa E todos requires_term casam E nenhum exclude_term casa.

    `sealed_only=True` (fonte secondhand-first, ex.: Enjoei) exige PROVA de lacre
    no título — sem isso, 0 candidatos. Fontes new-first usam o default (False)."""
    # Guards de fora-de-escopo ANTES do match: este repo é SELADO-only (Amazon/
    # OLX/ML buscam só selado; Liga navega categorias de selado). Logo um single
    # ("Eevee SVP 173 ... ETB ... Single Card") ou um acessório ("ETB ...
    # Acessórios", "Fichário ... Binder") que vem como ruído do marketplace casaria
    # o SKU do box e geraria margem fantasma. Fora do escopo → 0 candidatos.
    if looks_like_single_card(title) or looks_like_accessory(title):
        return []
    # Produto explicitamente ABERTO/USADO/INCOMPLETO nunca é compra de selado
    # (rejeição global, validada zero-regressão).
    if looks_used(title):
        return []
    # Fonte secondhand-first: "usado até provar lacre" — sem token de lacre, barra.
    if sealed_only and not looks_sealed(title):
        return []
    norm = normalize(title)
    candidates: list[Sku] = []
    for sku in registry:
        if not any(contains_term(norm, t) for t in sku.set_terms):
            continue
        if not any(contains_term(norm, t) for t in sku.type_terms):
            continue
        if sku.requires_terms and not all(contains_term(norm, t) for t in sku.requires_terms):
            continue
        if any(contains_term(norm, t) for t in sku.exclude_terms):
            continue
        candidates.append(sku)

    # Regra do guarda-chuva de era: se um SKU umbrella (set = nome de era, ex.:
    # "Mega Evolution") casou junto com um SKU de sub-set mais específico
    # (ex.: Perfect Order), o umbrella perde. Título "Mega Evolution - Perfect
    # Order Booster Bundle" → po-bundle, não meg-bundle. Genérico: cobre qualquer
    # sub-set da era sem enumerá-los em exclude_terms. Só age quando há mistura;
    # um anúncio do set-base puro ("Mega Evolution Booster Bundle") casa só o
    # umbrella e é preservado.
    specific = [c for c in candidates if not c.era_umbrella]
    if specific and len(specific) < len(candidates):
        candidates = specific
    return candidates


# --------------------------------------------------------------------------
# Cálculo de margem (convenção ROI: lucro sobre o capital investido)
# --------------------------------------------------------------------------
def _parse_price(raw) -> float:
    """Preço malformado de UM anúncio não derruba o run: vira 0.0 e é barrado como
    RED. Selado roda SEM piso (filters.min_brazil_price_brl: 0), então quem pega o
    0.0 é o zero-guard de compute_margin (`/ price_brl if price_brl else 0.0`) ->
    margem 0% < 30% -> RED `margem_total_abaixo_do_minimo` (nunca GREEN). Mesma
    degradação graciosa do qty e do SourceBlockedError."""
    try:
        return float(raw)
    except (TypeError, ValueError):
        return 0.0


def compute_margin(price_brl: float, us_usd: float, config: dict) -> dict:
    """Só margem BRUTA = (preço_US − preço_BR) / preço_BR. Custos operacionais
    (taxas, frete, 3PL, lote) ficam FORA do scanner — o operador calcula por
    fora. O scanner NÃO computa nem exibe margem líquida."""
    fx = config["currency"]["usd_brl"]

    us_brl = us_usd * fx
    gross_profit = us_brl - price_brl
    # Margem total do negócio: lucro sobre o capital de compra (convenção ROI).
    total_margin = gross_profit / price_brl if price_brl else 0.0
    # Quanto o BR está mais barato que a referência US (denominador = preço US).
    us_discount = gross_profit / us_brl if us_brl else 0.0

    return {
        "us_price_brl": round(us_brl, 2),
        "gross_profit_brl": round(gross_profit, 2),
        "total_margin_pct": round(total_margin, 4),
        "us_discount_pct": round(us_discount, 4),
    }


# --------------------------------------------------------------------------
# Classificação
# --------------------------------------------------------------------------
def classify(row: ScanRow, registry: list[Sku], us_reference: dict, config: dict) -> ScanRow:
    criteria = config["deal_criteria"]
    min_total = criteria["min_total_margin_pct"]
    min_price = config["filters"]["min_brazil_price_brl"]

    # Fontes secondhand-first (marketplace de usados, ex.: Enjoei) exigem prova
    # de lacre. Declaradas em scope.sealed_only_sources; demais = new-first.
    sealed_only_sources = {
        str(s).lower() for s in (config.get("scope", {}).get("sealed_only_sources") or [])
    }
    sealed_only = (row.source or "").lower() in sealed_only_sources

    candidates = match_listing(row.title_br, registry, sealed_only=sealed_only)
    row.match_candidates = [c.id for c in candidates]

    # --- sem match -------------------------------------------------------
    if not candidates:
        row.match_confidence = "NONE"
        row.deal_confidence = "RED"
        row.bucket = "rejected"
        norm = normalize(row.title_br)
        if looks_used(row.title_br):
            row.reject_reason = "produto_aberto_usado"
            row.main_risk = "Título sinaliza produto aberto/usado/incompleto — fora do escopo de selado"
        elif sealed_only and not looks_sealed(row.title_br):
            row.reject_reason = "lacre_nao_confirmado"
            row.main_risk = "Fonte de usados sem prova de lacre no título — não confirmável como selado"
        elif any(tok in norm.split() for tok in NON_EN_LANGUAGE_TOKENS):
            row.reject_reason = "idioma_nao_ingles"
            row.main_risk = "Produto não-inglês — sem liquidez no TCGPlayer"
        elif looks_like_single_card(row.title_br) or contains_term(norm, "carta"):
            row.reject_reason = "nao_e_selado"
            row.main_risk = "Parece carta avulsa, fora do escopo de selados"
        elif looks_like_accessory(row.title_br):
            row.reject_reason = "nao_e_selado"
            row.main_risk = "Parece acessório (binder/fichário/álbum/etc.), fora do escopo de selados"
        else:
            row.reject_reason = "sem_match_no_registry"
            row.main_risk = "Produto não está no registry curado de SKUs"
        row.recommended_action = "Ignorar (ou adicionar SKU ao registry se for selado válido)"
        return row

    # --- match ambíguo ---------------------------------------------------
    if len(candidates) > 1:
        row.match_confidence = "REVIEW"
        row.sku_name = " | ".join(c.name for c in candidates)
        row.product_type = candidates[0].product_type
        row.set_name = candidates[0].set_name
        row.deal_confidence = "YELLOW"
        row.bucket = "review_required"
        row.reject_reason = ""
        row.main_risk = (
            f"Anúncio casa com {len(candidates)} SKUs: {', '.join(row.match_candidates)}. "
            "Versão exata indefinida (ex.: ETB normal vs Pokémon Center)."
        )
        row.recommended_action = "Revisar o anúncio e identificar a versão exata antes de cotar"
        return row

    # --- match único (HIGH) ---------------------------------------------
    sku = candidates[0]
    row.match_confidence = "HIGH"
    row.sku_id = sku.id
    row.sku_name = sku.name
    row.product_type = sku.product_type
    row.set_name = sku.set_name

    # Piso de preço configurável. SELADO roda SEM piso (min_price=0, decisão do
    # operador 2026-06-27), então este ramo fica INATIVO por política — preço
    # 0/malformado é barrado adiante pelo zero-guard de compute_margin (margem 0%
    # < 30% -> RED). O ramo segue aqui de propósito: é o mecanismo genérico de
    # piso; NÃO reintroduzir o piso R$50 das cartas avulsas (vale só p/ singles).
    if min_price > 0 and row.price_brl < min_price:
        row.deal_confidence = "RED"
        row.bucket = "rejected"
        row.reject_reason = "abaixo_do_preco_minimo"
        row.main_risk = f"Preço R$ {row.price_brl:.2f} abaixo do mínimo de operação"
        row.recommended_action = "Ignorar"
        return row

    us_usd = us_reference.get(sku.id)
    if us_usd is None:
        # Nunca inventar preço.
        row.deal_confidence = "RED"
        row.bucket = "rejected"
        row.reject_reason = "sem_referencia_us"
        row.main_risk = "Sem preço de referência no TCGPlayer para este SKU"
        row.recommended_action = "Coletar referência TCGPlayer antes de avaliar"
        return row

    fin = compute_margin(row.price_brl, us_usd, config)
    row.us_price_usd = us_usd
    row.usd_brl = config["currency"]["usd_brl"]
    row.us_price_brl = fin["us_price_brl"]
    row.gross_profit_brl = fin["gross_profit_brl"]
    row.total_margin_pct = fin["total_margin_pct"]
    row.us_discount_pct = fin["us_discount_pct"]

    total_m = fin["total_margin_pct"]

    # Classificação por MARGEM BRUTA apenas. GREEN se a margem bruta atinge o
    # piso; senão RED. YELLOW é EXCLUSIVAMENTE match ambíguo (REVIEW, tratado
    # acima) — nunca por faixa de margem. Custos operacionais (frete, taxas,
    # lote) ficam FORA do scanner; o operador calcula por fora. Sem margem líquida.
    review_above = criteria.get("review_above_margin_pct")
    if total_m >= min_total and review_above is not None and total_m >= review_above:
        # Margem implausível p/ selado: forte sinal de produto/variante trocada
        # (acessório barato casando SKU caro — ex.: fichário avulso ~R$230 vs o
        # "151 Binder Collection" selado de US$240 -> 432%). NÃO vira GREEN; e
        # NÃO vira YELLOW (YELLOW é só match ambíguo, invariante). RED honesto
        # com motivo auditável, p/ o operador abrir e conferir o anúncio.
        row.deal_confidence = "RED"
        row.bucket = "rejected"
        row.reject_reason = "margem_anomala"
        row.main_risk = (
            f"Margem {total_m:.0%} alta demais p/ selado — provável produto/variante "
            "trocada (ex.: acessório/fichário avulso casando SKU caro). Verifique antes de comprar."
        )
        row.recommended_action = "Abrir o anúncio e confirmar que é o produto selado correto"
    elif total_m >= min_total:
        row.deal_confidence = "GREEN"
        row.bucket = "real_opportunities"
        row.main_risk = "Margem bruta — custo de frete/lote a cotar fora do scanner"
        row.recommended_action = "Validar estoque e cotar frete/lote"
    else:
        row.deal_confidence = "RED"
        row.bucket = "rejected"
        row.reject_reason = "margem_total_abaixo_do_minimo"
        row.main_risk = f"Margem total {total_m:.1%} abaixo do mínimo de {min_total:.0%}"
        row.recommended_action = "Ignorar"
    return row


# --------------------------------------------------------------------------
# Saída — CSV
# --------------------------------------------------------------------------
CSV_COLUMNS = [
    ("listing_id", "ID Anúncio"),
    ("title_br", "Título (BR)"),
    ("source", "Fonte"),
    ("seller", "Vendedor"),
    ("url", "URL"),
    ("price_brl", "Preço BR (R$)"),
    ("qty_avail", "Qtd disponível"),
    ("sku_id", "SKU"),
    ("sku_name", "Produto (canônico)"),
    ("product_type", "Tipo"),
    ("set_name", "Coleção"),
    ("us_price_usd", "Preço US (US$)"),
    ("usd_brl", "Câmbio USD/BRL"),
    ("us_price_brl", "Preço US (R$)"),
    ("gross_profit_brl", "Lucro bruto (R$)"),
    ("total_margin_pct", "Margem total %"),
    ("us_discount_pct", "Mais barato que US %"),
    ("match_confidence", "Confiança do match"),
    ("deal_confidence", "Confiança do deal"),
    ("main_risk", "Risco principal"),
    ("recommended_action", "Ação recomendada"),
    ("reject_reason", "Motivo de rejeição"),
]


def cell_value(row: ScanRow, key: str):
    val = getattr(row, key)
    if val is None:
        return ""
    if key in ("total_margin_pct", "us_discount_pct"):
        return f"{val * 100:.2f}"
    return val


def write_csv(rows: list[ScanRow], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow([label for _, label in CSV_COLUMNS])
        for row in rows:
            writer.writerow([cell_value(row, key) for key, _ in CSV_COLUMNS])


# --------------------------------------------------------------------------
# Saída — XLSX
# --------------------------------------------------------------------------
def write_xlsx(buckets: dict, config: dict, source_desc: str, path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [aviso] openpyxl não instalado — XLSX pulado (CSVs gerados normalmente).")
        return False

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    fills = {
        "GREEN": PatternFill("solid", fgColor="C6EFCE"),
        "YELLOW": PatternFill("solid", fgColor="FFEB9C"),
        "RED": PatternFill("solid", fgColor="F2DCDB"),
    }

    def add_sheet(name: str, rows: list[ScanRow]) -> None:
        ws = wb.create_sheet(name)
        ws.append([label for _, label in CSV_COLUMNS])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append([cell_value(row, key) for key, _ in CSV_COLUMNS])
            fill = fills.get(row.deal_confidence)
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill
        for idx, (key, label) in enumerate(CSV_COLUMNS, start=1):
            width = max(len(label) + 2, 14)
            if key in ("title_br", "sku_name", "main_risk", "recommended_action", "url"):
                width = 42
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"

    add_sheet("Real Opportunities", buckets["real_opportunities"])
    add_sheet("Review Required", buckets["review_required"])
    add_sheet("Rejected", buckets["rejected"])

    # Premissas — nunca esconder.
    ws = wb.create_sheet("Assumptions")
    ws.append(["Premissa", "Valor"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    assumptions = [
        ("Fonte dos anúncios", source_desc),
        ("Referência US", ", ".join(config["sources"]["usa"])),
        ("Câmbio USD/BRL", config["currency"]["usd_brl"]),
        ("Modo de câmbio", config["currency"]["mode"]),
        ("Preço mínimo BR (R$)", config["filters"]["min_brazil_price_brl"]),
        ("Margem bruta mínima (GREEN)", f"{config['deal_criteria']['min_total_margin_pct']:.0%}"),
        ("Custos operacionais", "fora do scanner (operador calcula por fora)"),
        ("Margem líquida", "não calculada — só margem bruta"),
    ]
    for label, value in assumptions:
        ws.append([label, value])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48

    # Summary
    ws = wb.create_sheet("Summary", 0)
    ws.append(["Bucket", "Qtd"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for label, key in (
        ("Real Opportunities", "real_opportunities"),
        ("Review Required", "review_required"),
        ("Rejected", "rejected"),
    ):
        ws.append([label, len(buckets[key])])
    ws.append(["TOTAL", sum(len(v) for v in buckets.values())])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14

    wb.save(path)
    return True


# --------------------------------------------------------------------------
# Orquestração
# --------------------------------------------------------------------------
def reference_age_days(captured_at: str | None) -> int | None:
    """Idade (em dias INTEIROS) da referência US a partir do `captured_at` do
    us_reference.json. None se ausente/ilegível (não força rebaixamento).

    `.days` trunca pra baixo: uma referência de 14,5 dias retorna 14. Logo o
    threshold `max_reference_age_days: N` tolera até quase N+1 dias — lenient por
    até 1 dia, nunca mais estrito. Imaterial no fluxo diário; relevante só se
    alguém setar N muito pequeno."""
    if not captured_at:
        return None
    try:
        dt = datetime.strptime(captured_at, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
    except (ValueError, TypeError):
        return None
    return (datetime.now(timezone.utc) - dt).days


def apply_freshness_downgrade(rows: list, ref_data: dict, config: dict) -> int | None:
    """Freshness guard FP-safe: referência US velha pode inflar margem em GREEN
    falso (o modo de falha histórico dos tins). Acima de `max_reference_age_days`,
    rebaixa CADA GREEN -> YELLOW (revisão manual), movendo-o p/ o bucket
    `review_required` com main_risk/recommended_action explicando o motivo.

    Nunca cria deal nem mexe em YELLOW/RED: só pede conferência do preço atual.
    Muta `rows` in-place. Retorna a idade (dias) da referência se ela estiver
    vencida, senão None (caller usa p/ logar). Caminho ÚNICO compartilhado entre
    o single-source (run) e o orquestrador (run_all_sources) — sem duplicar lógica.
    """
    ref_age = reference_age_days(ref_data.get("captured_at"))
    max_age = config.get("deal_criteria", {}).get("max_reference_age_days", 14)
    ref_stale = ref_age is not None and ref_age > max_age
    if not ref_stale:
        return None
    for row in rows:
        if row.deal_confidence == "GREEN":
            # Rebaixa GREEN -> YELLOW (revisão) quando a referência está velha:
            # o sinal de margem pode estar desatualizado. Move p/ o bucket de
            # revisão (review_required) p/ o operador conferir o TCGplayer atual.
            row.deal_confidence = "YELLOW"
            row.bucket = "review_required"
            row.main_risk = (
                f"Referência US defasada ({ref_age}d) — margem pode estar desatualizada; "
                "confira o preço TCGplayer atual antes de comprar."
            )
            row.recommended_action = "Refrescar referência (build_us_reference.py) e reconferir"
    return ref_age


def run(args: argparse.Namespace) -> int:
    config = load_yaml(Path(args.config), "config.yaml")
    fx_source = resolve_fx_rate(config)
    config["currency"]["_source"] = fx_source
    print(f"  [fx] cambio USD/BRL: {fx_source}")
    registry_data = load_yaml(Path(args.registry), "sku_registry.yaml")
    registry = build_registry(registry_data)

    try:
        listings, source_desc = load_listings(
            args.source, Path(args.mock), config, registry_raw=registry_data.get("skus", []),
        )
    except NotImplementedError as exc:
        print(f"ERRO: {exc}")
        return 2
    except SourceBlockedError as exc:
        # Fonte bloqueada por proteção EXTERNA (Cloudflare WAF / IP reputation).
        # O pipeline está OK; o site é que nega acesso. Condição não-fatal:
        # as outras fontes (--source amazon/liga) seguem operacionais.
        print()
        print("=" * 64)
        print(f"  FONTE BLOQUEADA (externo, não é bug do scanner)")
        print("=" * 64)
        print(f"  Fonte   : {exc.source}")
        print(f"  Motivo  : {exc.detail}")
        if exc.hint:
            print(f"  Nota    : {exc.hint}")
        print("=" * 64)
        return 0

    if not listings:
        # Fonte respondeu mas veio vazia (seletor quebrado, inventário vazio).
        # Isto É um problema real — não confundir com block externo acima.
        print("ERRO: nenhum anúncio carregado — fonte vazia ou quebrada.")
        return 1

    # Referência US (TCGPlayer via tcgcsv.com) — independente da fonte BR.
    # Rode sealed/build_us_reference.py para refrescar os preços.
    ref_data = load_json(SCRIPT_DIR / "data" / "us_reference.json", "data/us_reference.json")
    us_reference = ref_data.get("prices", {})

    rows: list[ScanRow] = []
    for item in listings:
        qty_raw = item.get("qty_avail")
        qty_avail: int | None
        try:
            qty_avail = int(qty_raw) if qty_raw is not None else None
        except (TypeError, ValueError):
            qty_avail = None
        row = ScanRow(
            listing_id=str(item.get("id", "")),
            title_br=item.get("title", ""),
            source=item.get("source", args.source),
            seller=item.get("seller", ""),
            url=item.get("url", ""),
            price_brl=_parse_price(item.get("price_brl", 0.0)),
            qty_avail=qty_avail,
        )
        classify(row, registry, us_reference, config)
        rows.append(row)

    # Freshness guard — referência US velha pode inflar margem em GREEN falso (o
    # modo de falha histórico dos tins). Acima da validade, rebaixa GREEN->YELLOW
    # (revisão manual). FP-safe: nunca cria deal, só pede conferência do preço atual.
    stale_age = apply_freshness_downgrade(rows, ref_data, config)
    if stale_age is not None:
        max_age = config.get("deal_criteria", {}).get("max_reference_age_days", 14)
        print(
            f"  [aviso] referência US tem {stale_age} dias (> {max_age}) — GREEN serão "
            f"rebaixados p/ YELLOW (conferência). Rode build_us_reference.py p/ refrescar."
        )

    buckets = {"real_opportunities": [], "review_required": [], "rejected": []}
    for row in rows:
        buckets[row.bucket].append(row)
    # Ordena cada balde por margem total desc — os maiores deltas saltam primeiro.
    for key in buckets:
        buckets[key].sort(
            key=lambda r: r.total_margin_pct if r.total_margin_pct is not None else -1,
            reverse=True,
        )

    # Diretório novo por run — cada scan é fresco, nada de misturar resultados.
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = SCRIPT_DIR / "results" / stamp
    out_dir.mkdir(parents=True, exist_ok=True)

    for key in buckets:
        write_csv(buckets[key], out_dir / f"{key}.csv")

    xlsx_path = out_dir / f"sealed_scan_{stamp}.xlsx"
    xlsx_ok = write_xlsx(buckets, config, source_desc, xlsx_path)

    print()
    print("=" * 64)
    print("  TCG SEALED ARBITRAGE SCANNER")
    print("=" * 64)
    print(f"  Fonte             : {source_desc}")
    print(f"  Referência US     : {', '.join(config['sources']['usa'])}")
    print(f"  Câmbio USD/BRL    : {config['currency']['usd_brl']:.4f}  [{config['currency'].get('_source', 'manual')}]")
    print(f"  Anúncios lidos    : {len(rows)}")
    print(f"  SKUs no registry  : {len(registry)}")
    print("-" * 64)
    print(f"  Oportunidades reais (GREEN) : {len(buckets['real_opportunities'])}")
    print(f"  Revisar manualmente         : {len(buckets['review_required'])}")
    print(f"  Rejeitados                  : {len(buckets['rejected'])}")
    print("-" * 64)
    for row in buckets["real_opportunities"]:
        print(f"  GREEN  {row.listing_id}  {row.sku_name}")
        print(f"         compra R$ {row.price_brl:.2f} | margem total "
              f"{row.total_margin_pct:.1%} | {row.us_discount_pct:.1%} mais barato que US")
    for row in buckets["review_required"]:
        print(f"  YELLOW {row.listing_id}  {row.title_br}  [{row.match_confidence}]")
    # Near-misses: HIGH-match rejected, sorted by margin desc. Mostra sempre
    # que GREEN ficar abaixo de 3 — ajuda a ver o que está perto do alvo.
    if len(buckets["real_opportunities"]) < 3:
        with_margin = [r for r in buckets["rejected"] if r.total_margin_pct is not None]
        if with_margin:
            print(f"  Top {min(5, len(with_margin))} matches HIGH rejeitados (near-misses, margem desc):")
            for row in with_margin[:5]:
                tag = "RED " if row.total_margin_pct < 0 else "RED+"
                print(f"  {tag}  {row.listing_id}  {row.sku_name}")
                print(f"         compra R$ {row.price_brl:.2f} | margem total "
                      f"{row.total_margin_pct:.1%} | {row.us_discount_pct:.1%} mais barato que US")
    print("-" * 64)
    print(f"  Resultados em: {out_dir}")
    if xlsx_ok:
        print(f"  XLSX         : {xlsx_path.name}")
    print("=" * 64)
    return 0


def main() -> None:
    from lib.console import harden_stdout
    harden_stdout()  # console Windows cp1252 quebra em títulos Liga/PT-BR
    parser = argparse.ArgumentParser(description="TCG Sealed Arbitrage Scanner (Brasil -> EUA)")
    parser.add_argument("--source", default="mock",
                        choices=["mock", "amazon", "olx", "mercadolivre", "liga"],
                        help="fonte dos anúncios (default: mock)")
    parser.add_argument("--config", default=str(SCRIPT_DIR / "config.yaml"),
                        help="caminho do config.yaml")
    parser.add_argument("--registry", default=str(SCRIPT_DIR / "sku_registry.yaml"),
                        help="caminho do sku_registry.yaml")
    parser.add_argument("--mock", default=str(SCRIPT_DIR / "mock_data" / "liga_listings.json"),
                        help="caminho do JSON de anúncios mockados")
    args = parser.parse_args()
    sys.exit(run(args))


if __name__ == "__main__":
    main()
